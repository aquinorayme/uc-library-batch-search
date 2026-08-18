#!/usr/bin/env python3
"""
check_uc_library.py

Look up each (Title, Subtitle, Author) row in a spreadsheet against UC
Berkeley's "UC Library Search" catalog (Primo VE) and record whether a
matching, UCB-owned record was found.
"""

import json
import re
import sys
import time
import requests
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------
API_URL = "https://search.library.berkeley.edu/primaws/rest/pub/pnxs"
VID = "01UCS_BER:UCB"
SCOPE = "DN_and_CI"
TAB = "Everything"
INST = "01UCS_BER"
UCB_ORG_CODE = "01UCS_BER"
ALMA_FALLBACK_URL_TEMPLATE = "https://na07.alma.exlibrisgroup.com/primaws/rest/pub/pnxs/L/{alma_id}?vid=" + VID

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "application/json",
}

COL_TITLE = "Title"
COL_SUBTITLE = "Subtitle"
COL_AUTHOR = "Author / Editor"
COL_RESULT = "UCLS"
COL_PRIORITY = "Priority"
COL_FORMAT = "Format"
COL_STATUS = "UCLS_Lookup_Result"
COL_MATCH_TITLE = "UCLS_Matched_Title"
COL_MATCH_AUTHOR = "UCLS_Matched_Author"
COL_MATCH_URL = "UCLS_Record_URL"
COL_MATCH_LOCATION = "UCLS_Location"

def normalize(s):
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def get_catalog_title_only(raw_title):
    return raw_title.split(" / ", 1)[0]

def author_tokens(s):
    if not s:
        return set()
    clean = s.split("$$Q")[0]
    return set(normalize(clean).split())

def extract_primary_author_for_query(author_field):
    if not author_field:
        return ""
    clean_a = re.sub(r'(?i)\b\(?eds?\.?\)?\b', '', author_field).lower()
    for fluff in ('introduction by', 'intro by', 'illustrations by',
                  'illustrated by', 'translated by', 'edited by'):
        if fluff in clean_a:
            clean_a = clean_a.split(fluff)[0]
    return re.split(r"&|,|;| and |\n", clean_a)[0].strip()

def check_ownership(delivery):
    if not delivery:
        return False, ""
    bestlocation = delivery.get("bestlocation") or {}
    if bestlocation.get("organization") == UCB_ORG_CODE:
        return True, bestlocation.get("mainLocation", "UCB Library")
    for holding in (delivery.get("holding") or []):
        if holding.get("organization") == UCB_ORG_CODE:
            return True, holding.get("mainLocation", "UCB Library")
    if delivery.get("electronicServices"):
        return True, "Online/Electronic"
    return False, ""

def search_book(title, subtitle, author, session, timeout=15):
    title_str = str(title or "").strip()
    sub_str = str(subtitle or "").strip()
    author_str = str(author or "").strip()

    title_str = re.sub(r'(?i)\(.*?[Ss]eries.*?\)', '', title_str).strip()
    sub_str = re.sub(r'(?i)\(.*?[Ss]eries.*?\)', '', sub_str).strip()

    if not title_str and not author_str:
        return {"status": "Skipped", "reason": "no title/author"}

    req_full_title = normalize(f"{title_str} : {sub_str}" if sub_str else title_str)
    req_author_tokens = author_tokens(extract_primary_author_for_query(author_str) or author_str)

    api_title = re.sub(r"[^a-zA-Z0-9\s]", " ", title_str)
    api_title = re.sub(r"\s+", " ", api_title).strip()
    api_author = re.sub(r"[^a-zA-Z0-9\s]", " ", extract_primary_author_for_query(author_str))
    api_author = re.sub(r"\s+", " ", api_author).strip()

    clauses = []
    if api_title:
        clauses.append(f"any,contains,{api_title},AND")
    if api_author:
        clauses.append(f"creator,contains,{api_author}")
    if not clauses:
        return {"status": "Skipped", "reason": "no valid query parameters"}

    params = {
        "q": ";".join(clauses),
        "vid": VID,
        "tab": TAB,
        "scope": SCOPE,
        "inst": INST,
        "lang": "en",
        "offset": "0",
        "limit": "50",
        "pcAvailability": "true",
    }

    try:
        resp = session.get(API_URL, params=params, headers=HEADERS, timeout=timeout)
        resp.raise_for_status()
        data = resp.json()
    except requests.exceptions.RequestException as e:
        return {"status": "Error", "reason": str(e)}
    except json.JSONDecodeError:
        return {"status": "Error", "reason": "non-JSON response"}

    docs = data.get("docs") or []
    if not docs:
        return {"status": "Not Found", "reason": "no results"}

    saw_exact_match_not_owned = False

    for doc in docs:
        pnx = doc.get("pnx") or {}
        display = pnx.get("display") or {}
        control = pnx.get("control") or {}

        cand_title_raw = " ".join(display.get("title") or [])
        cand_creator_raw = " ".join(display.get("creator") or [])
        cand_type = [t.lower() for t in (display.get("type") or [])]

        if not any("book" in t for t in cand_type):
            continue

        clean_cand_title = get_catalog_title_only(cand_title_raw)
        if normalize(clean_cand_title) != req_full_title:
            continue

        cand_author_tokens = author_tokens(cand_creator_raw)
        if req_author_tokens != cand_author_tokens and not (
            req_author_tokens and req_author_tokens.issubset(cand_author_tokens)
        ):
            continue

        record_id = doc.get("@id", "")
        rec_url = (
            f"https://search.library.berkeley.edu/discovery/fulldisplay?"
            f"docid={record_id.split('/')[-1] if record_id else ''}&vid={VID}"
        )

        delivery = doc.get("delivery") or {}
        is_owned, location_str = check_ownership(delivery)

        if not is_owned:
            record_ids = control.get("recordid") or []
            alma_id = next((rid for rid in record_ids if str(rid).startswith("alma")), None)
            if not alma_id and "alma" in record_id:
                alma_id = record_id.split("/")[-1]
            if alma_id:
                try:
                    f_resp = session.get(
                        ALMA_FALLBACK_URL_TEMPLATE.format(alma_id=alma_id),
                        headers=HEADERS, timeout=5,
                    )
                    if f_resp.status_code == 200:
                        f_delivery = (f_resp.json() or {}).get("delivery") or {}
                        is_owned, location_str = check_ownership(f_delivery)
                except requests.exceptions.RequestException:
                    pass

        if is_owned:
            return {
                "status": "Found",
                "matched_title": cand_title_raw,
                "matched_author": cand_creator_raw,
                "url": rec_url,
                "location": location_str,
            }
        else:
            saw_exact_match_not_owned = True

    if saw_exact_match_not_owned:
        return {"status": "Not Found (exact title/author match exists, but not a confirmed UCB holding)"}
    return {"status": "Not Found"}

def main():
    print("=========================================================")
    print("      UC Library Search - Configuration Wizard           ")
    print("=========================================================")
    
    # 1. Ask for input file
    input_xlsx = input("Enter Input Excel file [Current  (6).xlsx]: ").strip() or "Current  (6).xlsx"
    
    # 2. Ask for output file
    output_xlsx = input("Enter Output Excel file [final_validated_output.xlsx]: ").strip() or "final_validated_output.xlsx"
    
    # 3. Ask for Priority
    p_input = input("Target Priority (e.g., '1', '5', '1,2', or 'all') [5]: ").strip() or "5"
    allowed_priorities = [p.strip().lower() for p in p_input.split(",")]
    
    # 4. Ask for Format
    f_input = input("Target Format (e.g., 'book', 'monograph', or 'all') [book]: ").strip() or "book"
    allowed_formats = [f.strip().lower() for f in f_input.split(",")]
    
    # 5. Ask for Recheck rule
    r_input = input("Recheck rows that already have a result? (y/n) [n]: ").strip().lower()
    recheck_all = (r_input == 'y')
    
    print("=========================================================")
    print(f"Loading {input_xlsx}...")

    try:
        wb = openpyxl.load_workbook(input_xlsx)
    except Exception as e:
        print(f"ERROR: Could not open '{input_xlsx}'. {e}")
        sys.exit(1)

    ws = wb.active # Use the active/first sheet
    header = [c.value for c in ws[1]]
    col_idx = {name: i + 1 for i, name in enumerate(header) if name}

    required = [COL_TITLE, COL_AUTHOR]
    missing = [c for c in required if c not in col_idx]
    if missing:
        print(f"ERROR: missing expected column(s) {missing}")
        sys.exit(1)

    for new_col in (COL_STATUS, COL_MATCH_TITLE, COL_MATCH_AUTHOR, COL_MATCH_URL, COL_MATCH_LOCATION):
        if new_col not in col_idx:
            new_idx = ws.max_column + 1
            ws.cell(row=1, column=new_idx, value=new_col)
            col_idx[new_col] = new_idx

    session = requests.Session()
    processed = 0
    total_rows = ws.max_row
    
    print("\nStarting search process...\n")

    for row in range(2, total_rows + 1):
        title = ws.cell(row=row, column=col_idx[COL_TITLE]).value
        subtitle = ws.cell(row=row, column=col_idx[COL_SUBTITLE]).value if COL_SUBTITLE in col_idx else ""
        author = ws.cell(row=row, column=col_idx[COL_AUTHOR]).value

        existing_ucls = ws.cell(row=row, column=col_idx[COL_RESULT]).value if COL_RESULT in col_idx else None
        priority_val = ws.cell(row=row, column=col_idx[COL_PRIORITY]).value if COL_PRIORITY in col_idx else None
        format_val = ws.cell(row=row, column=col_idx[COL_FORMAT]).value if COL_FORMAT in col_idx else None
        already_done = ws.cell(row=row, column=col_idx[COL_STATUS]).value

        if not title and not author:
            continue

        # Configurable Priority Filter
        if "all" not in allowed_priorities and COL_PRIORITY in col_idx:
            p_val = str(priority_val).strip().replace('.0', '').lower()
            if p_val not in allowed_priorities:
                continue

        # Configurable Format Filter
        if "all" not in allowed_formats and COL_FORMAT in col_idx:
            f_val = str(format_val).strip().lower()
            if f_val not in allowed_formats:
                continue

        # Handle Rechecking logic
        if not recheck_all and str(existing_ucls).strip().lower() == "yes":
            continue
        if already_done and not recheck_all:
            continue

        result = search_book(title, subtitle, author, session)

        ws.cell(row=row, column=col_idx[COL_STATUS], value=result.get("status", "Error"))
        ws.cell(row=row, column=col_idx[COL_MATCH_TITLE], value=result.get("matched_title", ""))
        ws.cell(row=row, column=col_idx[COL_MATCH_AUTHOR], value=result.get("matched_author", ""))
        ws.cell(row=row, column=col_idx[COL_MATCH_URL], value=result.get("url", ""))
        ws.cell(row=row, column=col_idx[COL_MATCH_LOCATION], value=result.get("location", ""))

        print(f"Row {row}: '{title}' -> {result.get('status')}")

        processed += 1
        if processed % 20 == 0:
            wb.save(output_xlsx)

        time.sleep(1.0) # Respectful delay to prevent API throttling

    # --- AUTO FORMATTING THE OUTPUT ---
    print("\nFormatting output columns for readability...")
    for col_name in [COL_STATUS, COL_MATCH_TITLE, COL_MATCH_AUTHOR, COL_MATCH_URL, COL_MATCH_LOCATION]:
        if col_name in col_idx:
            c_idx = col_idx[col_name]
            letter = get_column_letter(c_idx)
            ws.column_dimensions[letter].width = 45 # Expand width
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=c_idx).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_xlsx)
    print(f"\nDone. Processed {processed} rows. Saved to '{output_xlsx}'")

if __name__ == "__main__":
    main()
